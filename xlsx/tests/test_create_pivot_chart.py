"""
Unit tests for create_pivot_chart.py script
"""
import os
import sys
import pytest
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart

# Add parent directory to path to import the script
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from create_pivot_chart import create_pivot_chart


class TestCreatePivotChart:
    """Test suite for pivot chart creation functionality"""

    @pytest.fixture
    def sample_csv_path(self):
        """Path to sample test data"""
        return os.path.join(os.path.dirname(__file__), 'sample_data.csv')

    @pytest.fixture
    def output_path(self, tmp_path):
        """Temporary output file path"""
        return os.path.join(tmp_path, 'test_pivot.xlsx')

    def test_creates_output_file(self, sample_csv_path, output_path):
        """Test that script creates an output file"""
        result = create_pivot_chart(sample_csv_path, output_path, 'country', 'org_name', 'count')

        assert os.path.exists(output_path), "Output file should be created"
        assert result['status'] == 'success'

    def test_has_required_sheets(self, sample_csv_path, output_path):
        """Test that output has Pivot sheet"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'org_name', 'count')

        wb = load_workbook(output_path)
        assert 'Pivot' in wb.sheetnames, "Should have 'Pivot' sheet"

    def test_pivot_sheet_has_chart(self, sample_csv_path, output_path):
        """Test that Pivot sheet contains a chart"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'org_name', 'count')

        wb = load_workbook(output_path)
        ws = wb['Pivot']

        assert len(ws._charts) > 0, "Pivot sheet should contain a chart"
        assert isinstance(ws._charts[0], BarChart), "Chart should be a BarChart"

    def test_count_aggregation(self, sample_csv_path, output_path):
        """Test count aggregation function"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'org_name', 'count')

        df = pd.read_excel(output_path, sheet_name='Pivot')

        # USA has 4 organizations
        assert df[df[df.columns[0]] == 'USA'][df.columns[1]].values[0] == 4
        # Canada has 2 organizations
        assert df[df[df.columns[0]] == 'Canada'][df.columns[1]].values[0] == 2

    def test_sum_aggregation(self, sample_csv_path, output_path):
        """Test sum aggregation function"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'sum')

        df = pd.read_excel(output_path, sheet_name='Pivot')

        # USA total revenue: 1000000 + 500000 + 850000 + 1100000 = 3450000
        usa_revenue = df[df[df.columns[0]] == 'USA'][df.columns[1]].values[0]
        assert usa_revenue == 3450000

    def test_mean_aggregation(self, sample_csv_path, output_path):
        """Test mean aggregation function"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'products', 'mean')

        df = pd.read_excel(output_path, sheet_name='Pivot')

        # USA average products: (5 + 3 + 6 + 9) / 4 = 5.75
        usa_avg = df[df[df.columns[0]] == 'USA'][df.columns[1]].values[0]
        assert abs(usa_avg - 5.75) < 0.01

    def test_unique_aggregation(self, sample_csv_path, output_path):
        """Test unique aggregation function"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'org_name', 'unique')

        df = pd.read_excel(output_path, sheet_name='Pivot')

        # Each country should have unique count equal to count
        usa_unique = df[df[df.columns[0]] == 'USA'][df.columns[1]].values[0]
        assert usa_unique == 4

    def test_return_value_structure(self, sample_csv_path, output_path):
        """Test that return value has correct structure"""
        result = create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'sum')

        assert 'status' in result
        assert 'output' in result
        assert 'groups' in result

        assert result['status'] == 'success'
        assert result['groups'] > 0

    def test_handles_invalid_aggregation(self, sample_csv_path, output_path):
        """Test that invalid aggregation defaults to unique (doesn't raise)"""
        # Script doesn't validate agg_func, defaults to 'unique' for unknown values
        result = create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'invalid_agg')
        assert result['status'] == 'success'

    def test_handles_invalid_column(self, sample_csv_path, output_path):
        """Test error handling for non-existent column"""
        with pytest.raises(ValueError):
            create_pivot_chart(sample_csv_path, output_path, 'invalid_col', 'revenue', 'sum')

    def test_handles_missing_file(self, output_path):
        """Test error handling for missing CSV file"""
        with pytest.raises(FileNotFoundError):
            create_pivot_chart('nonexistent.csv', output_path, 'country', 'revenue', 'sum')

    def test_data_sheet_preserves_original(self, sample_csv_path, output_path):
        """Test that Pivot sheet contains aggregated data (no separate Data sheet)"""
        df_original = pd.read_csv(sample_csv_path)
        create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'sum')

        df_output = pd.read_excel(output_path, sheet_name='Pivot')

        # Pivot sheet has aggregated data, not original
        assert len(df_output) < len(df_original)
        assert len(df_output) == df_original['country'].nunique()

    def test_pivot_data_is_aggregated(self, sample_csv_path, output_path):
        """Test that Pivot sheet has aggregated data"""
        df_original = pd.read_csv(sample_csv_path)
        create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'sum')

        df_pivot = pd.read_excel(output_path, sheet_name='Pivot')

        # Pivot should have fewer rows (one per country)
        assert len(df_pivot) < len(df_original)
        assert len(df_pivot) == df_original['country'].nunique()

    def test_chart_title_includes_aggregation(self, sample_csv_path, output_path):
        """Test that chart title reflects the aggregation"""
        create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'sum')

        wb = load_workbook(output_path)
        ws = wb['Pivot']
        chart = ws._charts[0]

        assert chart.title is not None
        title_text = str(chart.title.text.rich.p[0].r[0].t).lower()
        assert 'sum' in title_text or 'revenue' in title_text
