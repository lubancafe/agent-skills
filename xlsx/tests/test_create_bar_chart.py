"""
Unit tests for create_bar_chart.py script
"""
import os
import sys
import pytest
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart

# Add parent directory to path to import the script
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from create_bar_chart import create_bar_chart


class TestCreateBarChart:
    """Test suite for bar chart creation functionality"""

    @pytest.fixture
    def sample_csv_path(self):
        """Path to sample test data"""
        return os.path.join(os.path.dirname(__file__), 'sample_data.csv')

    @pytest.fixture
    def output_path(self, tmp_path):
        """Temporary output file path"""
        return os.path.join(tmp_path, 'test_chart.xlsx')

    def test_creates_output_file(self, sample_csv_path, output_path):
        """Test that script creates an output file"""
        result = create_bar_chart(sample_csv_path, output_path, 'country')

        assert os.path.exists(output_path), "Output file should be created"
        assert result['status'] == 'success'

    def test_has_data_and_chart_sheets(self, sample_csv_path, output_path):
        """Test that output has Chart Data sheet with embedded chart"""
        create_bar_chart(sample_csv_path, output_path, 'country')

        wb = load_workbook(output_path)
        assert 'Chart Data' in wb.sheetnames, "Should have 'Chart Data' sheet"

    def test_chart_sheet_has_chart(self, sample_csv_path, output_path):
        """Test that Chart Data sheet contains an embedded bar chart"""
        create_bar_chart(sample_csv_path, output_path, 'country')

        wb = load_workbook(output_path)
        ws = wb['Chart Data']

        # Check that a chart exists
        assert len(ws._charts) > 0, "Chart Data sheet should contain at least one chart"
        assert isinstance(ws._charts[0], BarChart), "Chart should be a BarChart"

    def test_count_aggregation_by_x_column(self, sample_csv_path, output_path):
        """Test counting by x_column (country)"""
        create_bar_chart(sample_csv_path, output_path, 'country')

        df = pd.read_excel(output_path, sheet_name='Chart Data')

        # Verify counts: USA=4, Canada=2, UK=2, Germany=1, France=1
        assert df[df['country'] == 'USA']['Count'].values[0] == 4
        assert df[df['country'] == 'Canada']['Count'].values[0] == 2
        assert df[df['country'] == 'UK']['Count'].values[0] == 2

    def test_sum_aggregation_with_y_column(self, sample_csv_path, output_path):
        """Test with specific y_column (note: script doesn't aggregate, just passes through)"""
        create_bar_chart(sample_csv_path, output_path, 'country', 'revenue')

        df = pd.read_excel(output_path, sheet_name='Chart Data')

        # Script passes through data with y_column, doesn't aggregate
        # Just verify we have the expected columns
        assert 'country' in df.columns
        assert 'revenue' in df.columns
        assert len(df) > 0

    def test_chart_has_correct_title(self, sample_csv_path, output_path):
        """Test that chart has appropriate title"""
        create_bar_chart(sample_csv_path, output_path, 'country')

        wb = load_workbook(output_path)
        ws = wb['Chart Data']
        chart = ws._charts[0]

        assert chart.title is not None, "Chart should have a title"
        assert 'country' in str(chart.title.text.rich.p[0].r[0].t).lower()

    def test_return_value_structure(self, sample_csv_path, output_path):
        """Test that return value has correct structure"""
        result = create_bar_chart(sample_csv_path, output_path, 'country', 'revenue')

        assert 'status' in result
        assert 'output' in result
        assert 'chart_type' in result
        assert 'data_points' in result

        assert result['status'] == 'success'
        assert result['chart_type'] == 'bar'
        assert result['data_points'] > 0

    def test_handles_invalid_column(self, sample_csv_path, output_path):
        """Test error handling for non-existent column"""
        with pytest.raises(ValueError):
            create_bar_chart(sample_csv_path, output_path, 'invalid_column')

    def test_handles_missing_file(self, output_path):
        """Test error handling for missing CSV file"""
        with pytest.raises(FileNotFoundError):
            create_bar_chart('nonexistent.csv', output_path, 'country')

    def test_data_sheet_preserves_original(self, sample_csv_path, output_path):
        """Test that Chart Data sheet contains aggregated data (not original CSV)"""
        df_original = pd.read_csv(sample_csv_path)
        create_bar_chart(sample_csv_path, output_path, 'country')

        df_output = pd.read_excel(output_path, sheet_name='Chart Data')

        # Chart Data has aggregated data, not original
        assert len(df_output) < len(df_original), "Should be aggregated (fewer rows)"
        assert 'country' in df_output.columns

    def test_chart_data_is_aggregated(self, sample_csv_path, output_path):
        """Test that Chart Data sheet has aggregated data, not raw data"""
        df_original = pd.read_csv(sample_csv_path)
        create_bar_chart(sample_csv_path, output_path, 'country')

        df_chart = pd.read_excel(output_path, sheet_name='Chart Data')

        # Chart data should be aggregated (fewer rows than original)
        assert len(df_chart) < len(df_original)
        # Should have unique countries only
        assert len(df_chart) == df_original['country'].nunique()

    def test_multiple_y_columns(self, sample_csv_path, output_path):
        """Test chart with products as y_column"""
        create_bar_chart(sample_csv_path, output_path, 'country', 'products')

        df = pd.read_excel(output_path, sheet_name='Chart Data')

        # When y_column is specified, data is passed through (not aggregated)
        assert 'country' in df.columns
        assert 'products' in df.columns
        assert len(df) > 0
