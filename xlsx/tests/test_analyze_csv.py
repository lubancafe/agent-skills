"""
Unit tests for analyze_csv.py script
"""
import os
import sys
import pytest
import pandas as pd
from openpyxl import load_workbook

# Add parent directory to path to import the script
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from analyze_csv import analyze_csv


class TestAnalyzeCsv:
    """Test suite for CSV analysis functionality"""

    @pytest.fixture
    def sample_csv_path(self):
        """Path to sample test data"""
        return os.path.join(os.path.dirname(__file__), 'sample_data.csv')

    @pytest.fixture
    def output_path(self, tmp_path):
        """Temporary output file path"""
        return os.path.join(tmp_path, 'test_output.xlsx')

    def test_analyze_csv_creates_file(self, sample_csv_path, output_path):
        """Test that analyze_csv creates an output file"""
        result = analyze_csv(sample_csv_path, output_path)

        assert os.path.exists(output_path), "Output file should be created"
        assert result['status'] == 'success'
        assert result['output'] == output_path

    def test_analyze_csv_has_correct_sheets(self, sample_csv_path, output_path):
        """Test that output has Data and Summary sheets"""
        analyze_csv(sample_csv_path, output_path)

        wb = load_workbook(output_path)
        assert 'Data' in wb.sheetnames, "Should have 'Data' sheet"
        assert 'Summary' in wb.sheetnames, "Should have 'Summary' sheet"

    def test_data_sheet_has_all_rows(self, sample_csv_path, output_path):
        """Test that Data sheet contains all CSV rows"""
        df_original = pd.read_csv(sample_csv_path)
        analyze_csv(sample_csv_path, output_path)

        df_output = pd.read_excel(output_path, sheet_name='Data')

        assert len(df_output) == len(df_original), "Should have same number of rows"
        assert list(df_output.columns) == list(df_original.columns), "Should have same columns"

    def test_data_sheet_header_formatting(self, sample_csv_path, output_path):
        """Test that header row has proper formatting"""
        analyze_csv(sample_csv_path, output_path)

        wb = load_workbook(output_path)
        ws = wb['Data']

        # Check first row (header) formatting
        for cell in ws[1]:
            assert cell.font.bold == True, "Header should be bold"
            # RGB color can be either FFFFFFFF or 00FFFFFF depending on openpyxl version
            assert cell.font.color.rgb in ['FFFFFFFF', '00FFFFFF'], "Header text should be white"
            assert cell.fill.start_color.rgb in ['FF4472C4', '004472C4'], "Header background should be blue"

    def test_summary_sheet_has_statistics(self, sample_csv_path, output_path):
        """Test that Summary sheet contains statistics"""
        analyze_csv(sample_csv_path, output_path)

        wb = load_workbook(output_path)
        ws = wb['Summary']

        # Check that summary has content
        assert ws['A1'].value is not None, "Summary should have content in A1"

    def test_return_value_structure(self, sample_csv_path, output_path):
        """Test that return value has correct structure"""
        result = analyze_csv(sample_csv_path, output_path)

        assert 'status' in result
        assert 'output' in result
        assert 'rows' in result
        assert 'columns' in result

        assert result['status'] == 'success'
        assert result['rows'] == 10  # sample_data.csv has 10 data rows
        assert result['columns'] == 4  # country, org_name, products, revenue

    def test_handles_missing_file(self, output_path):
        """Test error handling for missing CSV file"""
        with pytest.raises(FileNotFoundError):
            analyze_csv('nonexistent.csv', output_path)

    def test_output_is_valid_excel(self, sample_csv_path, output_path):
        """Test that output can be opened by openpyxl"""
        analyze_csv(sample_csv_path, output_path)

        # Should not raise exception
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) >= 2
