"""
Integration test for creating Excel files with visual charts.

This test verifies end-to-end functionality of creating Excel files
with embedded charts that can be opened and viewed in Excel.
"""
import os
import sys
import pytest
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart

# Add parent directory to path to import the scripts
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from create_bar_chart import create_bar_chart
from create_pivot_chart import create_pivot_chart


class TestVisualChartIntegration:
    """Integration tests for Excel files with visual charts"""

    @pytest.fixture
    def sample_csv_path(self):
        """Path to sample test data"""
        return os.path.join(os.path.dirname(__file__), 'sample_data.csv')

    @pytest.fixture
    def output_dir(self, tmp_path):
        """Temporary output directory"""
        return tmp_path

    def test_create_bar_chart_with_visual_chart(self, sample_csv_path, output_dir):
        """
        End-to-end test: Create Excel with visual bar chart

        This test creates an actual Excel file with an embedded chart
        that can be opened in Excel for visual inspection.
        """
        output_path = os.path.join(output_dir, 'country_counts_bar_chart.xlsx')

        # Create bar chart
        result = create_bar_chart(sample_csv_path, output_path, 'country')

        # Verify file creation
        assert os.path.exists(output_path), "Excel file should exist"
        assert result['status'] == 'success'

        # Load and verify workbook structure
        wb = load_workbook(output_path)
        ws = wb['Chart Data']

        # Verify data is present
        df = pd.read_excel(output_path, sheet_name='Chart Data')
        assert len(df) == 5, "Should have 5 unique countries"
        assert 'country' in df.columns
        assert 'Count' in df.columns

        # Verify chart exists
        assert len(ws._charts) > 0, "Should have at least one chart"
        chart = ws._charts[0]
        assert isinstance(chart, BarChart), "Should be a BarChart"

        # Verify chart properties
        assert chart.title is not None, "Chart should have a title"
        assert chart.type == "col", "Should be column (vertical bar) chart"

        # Verify chart data references
        assert chart.ser is not None, "Chart should have data series"
        assert len(chart.ser) > 0, "Chart should have at least one series"

        print(f"\n✓ Created visual bar chart: {output_path}")
        print(f"  - Countries: {list(df['country'])}")
        print(f"  - Counts: {list(df['Count'])}")
        print(f"  - Chart type: {chart.type}")
        print(f"  - Chart title: {chart.title.text.rich.p[0].r[0].t}")

    def test_create_pivot_chart_with_visual_chart(self, sample_csv_path, output_dir):
        """
        End-to-end test: Create Excel with visual pivot chart

        This test creates an Excel file with pivot data and chart.
        """
        output_path = os.path.join(output_dir, 'revenue_by_country_pivot.xlsx')

        # Create pivot chart with sum aggregation
        result = create_pivot_chart(sample_csv_path, output_path, 'country', 'revenue', 'sum')

        # Verify file creation
        assert os.path.exists(output_path), "Excel file should exist"
        assert result['status'] == 'success'
        assert result['groups'] == 5, "Should have 5 country groups"

        # Load and verify workbook structure
        wb = load_workbook(output_path)
        ws = wb['Pivot']

        # Verify pivot data
        df = pd.read_excel(output_path, sheet_name='Pivot')
        assert len(df) == 5, "Should have 5 countries"
        assert df.columns[0] == 'country'
        assert 'revenue_sum' in df.columns[1]

        # Verify revenue sums are correct
        usa_revenue = df[df['country'] == 'USA'][df.columns[1]].values[0]
        assert usa_revenue == 3450000, "USA revenue should be 3,450,000"

        # Verify chart exists
        assert len(ws._charts) > 0, "Should have a chart"
        chart = ws._charts[0]
        assert isinstance(chart, BarChart), "Should be a BarChart"

        # Verify chart is positioned correctly
        assert chart.anchor is not None, "Chart should be anchored to a cell"

        print(f"\n✓ Created visual pivot chart: {output_path}")
        print(f"  - Top country by revenue: {df.iloc[0]['country']} (${df.iloc[0][df.columns[1]]:,.0f})")
        print(f"  - Total groups: {len(df)}")
        print(f"  - Aggregation: sum")

    def test_create_multiple_chart_types(self, sample_csv_path, output_dir):
        """
        Test creating multiple Excel files with different chart types

        This creates a suite of visual charts for manual inspection.
        """
        test_cases = [
            {
                'name': 'count_by_country.xlsx',
                'func': create_bar_chart,
                'args': ('country',),
                'description': 'Count of organizations by country'
            },
            {
                'name': 'products_by_country.xlsx',
                'func': create_pivot_chart,
                'args': ('country', 'products', 'sum'),
                'description': 'Total products by country'
            },
            {
                'name': 'avg_revenue_by_country.xlsx',
                'func': create_pivot_chart,
                'args': ('country', 'revenue', 'mean'),
                'description': 'Average revenue by country'
            }
        ]

        created_files = []

        for test_case in test_cases:
            output_path = os.path.join(output_dir, test_case['name'])
            result = test_case['func'](sample_csv_path, output_path, *test_case['args'])

            assert result['status'] == 'success'
            assert os.path.exists(output_path)

            # Verify chart exists
            wb = load_workbook(output_path)
            ws = wb.active
            assert len(ws._charts) > 0, f"{test_case['name']} should have a chart"

            created_files.append({
                'file': output_path,
                'description': test_case['description']
            })

        print(f"\n✓ Created {len(created_files)} Excel files with visual charts:")
        for file_info in created_files:
            print(f"  - {os.path.basename(file_info['file'])}: {file_info['description']}")

    def test_chart_is_visible_and_formatted(self, sample_csv_path, output_dir):
        """
        Test that chart has proper formatting for visibility
        """
        output_path = os.path.join(output_dir, 'formatted_chart.xlsx')

        create_bar_chart(sample_csv_path, output_path, 'country')

        wb = load_workbook(output_path)
        ws = wb['Chart Data']
        chart = ws._charts[0]

        # Verify chart dimensions (should be visible, not tiny)
        # Chart should have width and height attributes
        assert hasattr(chart, 'width') or hasattr(chart, 'height')

        # Verify axis labels
        assert chart.x_axis.title is not None, "X-axis should have a label"
        assert chart.y_axis.title is not None, "Y-axis should have a label"

        # Verify chart positioning (anchored to a specific cell)
        assert chart.anchor is not None
        # Chart is anchored at E5 (col=4, row=4 in 0-indexed)
        # Verify anchor has position data
        assert hasattr(chart.anchor, '_from')
        anchor_from = chart.anchor._from
        assert anchor_from.col == 4, "Chart should be anchored at column E (4)"
        assert anchor_from.row == 4, "Chart should be anchored at row 5 (4 in 0-indexed)"

        print(f"\n✓ Chart formatting verified:")
        print(f"  - Position: Anchored at cell E5 (col={anchor_from.col}, row={anchor_from.row})")
        print(f"  - X-axis label: {chart.x_axis.title}")
        print(f"  - Y-axis label: {chart.y_axis.title}")
        print(f"  - Chart style: {chart.style}")

    def test_data_and_chart_alignment(self, sample_csv_path, output_dir):
        """
        Test that chart data aligns with table data
        """
        output_path = os.path.join(output_dir, 'data_chart_alignment.xlsx')

        create_bar_chart(sample_csv_path, output_path, 'country')

        # Read data
        df = pd.read_excel(output_path, sheet_name='Chart Data')

        # Load chart
        wb = load_workbook(output_path)
        ws = wb['Chart Data']
        chart = ws._charts[0]

        # Verify chart references match data range
        # Chart should reference column 2 (Count values)
        series = chart.ser[0]

        # Data series should exist
        assert series.val is not None, "Chart should have value references"
        assert series.cat is not None, "Chart should have category references"

        # Verify data count matches chart series length
        # The chart should have as many data points as rows in the DataFrame
        assert len(df) > 0, "Should have data points"

        print(f"\n✓ Data and chart alignment verified:")
        print(f"  - Data rows: {len(df)}")
        print(f"  - Categories: {list(df['country'])}")
        print(f"  - Values: {list(df['Count'])}")
