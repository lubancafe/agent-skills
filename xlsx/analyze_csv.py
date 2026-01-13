"""
Analyze CSV and generate summary statistics.

Usage:
    python analyze_csv.py <csv_path> <output_xlsx>

Example:
    python analyze_csv.py test.csv analysis.xlsx
"""
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def analyze_csv(csv_path, output_path):
    # Read CSV
    df = pd.read_csv(csv_path)

    # Create workbook
    wb = Workbook()

    # Sheet 1: Raw Data
    ws_data = wb.active
    ws_data.title = "Data"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    # Format header
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Sheet 2: Summary Statistics
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = 'Column'
    ws_summary['B1'] = 'Count'
    ws_summary['C1'] = 'Unique Values'
    ws_summary['D1'] = 'Most Common'

    # Format header
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

    # Add statistics
    row = 2
    for col in df.columns:
        ws_summary[f'A{row}'] = col
        ws_summary[f'B{row}'] = len(df[col].dropna())
        ws_summary[f'C{row}'] = df[col].nunique()
        if df[col].dtype == 'object':
            ws_summary[f'D{row}'] = df[col].mode()[0] if len(df[col].mode()) > 0 else 'N/A'
        row += 1

    # Adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(output_path)
    print(f"Analysis saved to {output_path}")
    return {"status": "success", "output": output_path, "rows": len(df), "columns": len(df.columns)}

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python analyze_csv.py <csv_path> <output_xlsx>")
        sys.exit(1)

    csv_path = sys.argv[1]
    output_path = sys.argv[2]
    result = analyze_csv(csv_path, output_path)
    print(f"Result: {result}")
