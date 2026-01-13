"""
Create bar chart from CSV data.

Usage:
    python create_bar_chart.py <csv_path> <output_xlsx> <x_column> [y_column]

Example:
    python create_bar_chart.py test.csv chart.xlsx country
    python create_bar_chart.py test.csv chart.xlsx org_name products
"""
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def create_bar_chart(csv_path, output_path, x_column, y_column=None):
    # Read CSV
    df = pd.read_csv(csv_path)

    if x_column not in df.columns:
        raise ValueError(f"Column '{x_column}' not found in CSV")

    # Prepare data for chart
    if y_column:
        # Use specified y column
        if y_column not in df.columns:
            raise ValueError(f"Column '{y_column}' not found in CSV")
        chart_data = df[[x_column, y_column]].copy()
    else:
        # Count occurrences of x_column values
        chart_data = df[x_column].value_counts().reset_index()
        chart_data.columns = [x_column, 'Count']

    # Sort by value
    chart_data = chart_data.sort_values(chart_data.columns[1], ascending=False)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart Data"

    # Write data
    for r in dataframe_to_rows(chart_data, index=False, header=True):
        ws.append(r)

    # Format header
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"{chart_data.columns[1]} by {x_column}"
    chart.y_axis.title = chart_data.columns[1]
    chart.x_axis.title = x_column

    # Data references
    data = Reference(ws, min_col=2, min_row=1, max_row=len(chart_data) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(chart_data) + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4

    # Add chart to sheet
    ws.add_chart(chart, "E5")

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    wb.save(output_path)
    print(f"Bar chart saved to {output_path}")
    return {"status": "success", "output": output_path, "chart_type": "bar", "data_points": len(chart_data)}

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: python create_bar_chart.py <csv_path> <output_xlsx> <x_column> [y_column]")
        sys.exit(1)

    csv_path = sys.argv[1]
    output_path = sys.argv[2]
    x_column = sys.argv[3]
    y_column = sys.argv[4] if len(sys.argv) > 4 else None

    result = create_bar_chart(csv_path, output_path, x_column, y_column)
    print(f"Result: {result}")
