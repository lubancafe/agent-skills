"""
Create pivot table with chart from CSV data.

Usage:
    python create_pivot_chart.py <csv_path> <output_xlsx> <index_col> <value_col> <agg_func>

Example:
    python create_pivot_chart.py test.csv pivot.xlsx country org_name count
    python create_pivot_chart.py test.csv pivot.xlsx country products count
"""
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def create_pivot_chart(csv_path, output_path, index_col, value_col, agg_func='count'):
    # Read CSV
    df = pd.read_csv(csv_path)

    if index_col not in df.columns:
        raise ValueError(f"Column '{index_col}' not found in CSV")
    if value_col not in df.columns:
        raise ValueError(f"Column '{value_col}' not found in CSV")

    # Create pivot
    if agg_func == 'count':
        pivot = df.groupby(index_col)[value_col].count().reset_index()
        pivot.columns = [index_col, f'{value_col}_count']
    elif agg_func == 'sum':
        pivot = df.groupby(index_col)[value_col].sum().reset_index()
        pivot.columns = [index_col, f'{value_col}_sum']
    elif agg_func == 'mean':
        pivot = df.groupby(index_col)[value_col].mean().reset_index()
        pivot.columns = [index_col, f'{value_col}_avg']
    else:
        pivot = df.groupby(index_col)[value_col].nunique().reset_index()
        pivot.columns = [index_col, f'{value_col}_unique']

    # Sort
    pivot = pivot.sort_values(pivot.columns[1], ascending=False)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot"

    # Write data
    for r in dataframe_to_rows(pivot, index=False, header=True):
        ws.append(r)

    # Format header
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

    # Create chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = f"{pivot.columns[1]} by {index_col}"
    chart.y_axis.title = pivot.columns[1]
    chart.x_axis.title = index_col

    data = Reference(ws, min_col=2, min_row=1, max_row=len(pivot) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(pivot) + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "E5")

    # Adjust columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    wb.save(output_path)
    print(f"Pivot chart saved to {output_path}")
    return {"status": "success", "output": output_path, "groups": len(pivot)}

if __name__ == '__main__':
    if len(sys.argv) < 6:
        print("Usage: python create_pivot_chart.py <csv_path> <output_xlsx> <index_col> <value_col> <agg_func>")
        print("  agg_func: count, sum, mean, unique")
        sys.exit(1)

    csv_path = sys.argv[1]
    output_path = sys.argv[2]
    index_col = sys.argv[3]
    value_col = sys.argv[4]
    agg_func = sys.argv[5]

    result = create_pivot_chart(csv_path, output_path, index_col, value_col, agg_func)
    print(f"Result: {result}")
